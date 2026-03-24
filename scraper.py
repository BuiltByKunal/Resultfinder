import argparse
import csv
import re
import sys
from pathlib import Path
from typing import Dict, Optional, Tuple

import requests
from bs4 import BeautifulSoup

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

BASE_URL = "https://result.mdu.ac.in/postexam/result.aspx"
DEFAULT_OUTPUT = "result.html"
DEFAULT_RESULTS_DIR = "results"
DEFAULT_SUMMARY = "results/index.html"

def _collect_form_data(soup: BeautifulSoup) -> Dict[str, str]:
    form = soup.find("form")
    container = form if form is not None else soup
    payload: Dict[str, str] = {};

    for field in container.find_all("input"):
        name = field.get("name") or field.get("id")
        if not name:
            continue;
        payload[name] = field.get("value", "")

    return payload

def _find_text_fields(soup: BeautifulSoup) -> Dict[str, Optional[str]]:
    form = soup.find("form")
    container = form if form is not None else soup

    text_inputs = [
        field
        for field in container.find_all("input")
        if field.get("type", "text").lower() in {"text", ""}
    ]

    reg_field = None;
    roll_field = None;

    for field in text_inputs:
        key = field.get("name") or field.get("id") or ""
        if "reg" in key.lower():
            reg_field = key;
            break;

    for field in text_inputs:
        key = field.get("name") or field.get("id") or ""
        if "roll" in key.lower():
            roll_field = key;
            break;

    if not reg_field and text_inputs:
        reg_field = text_inputs[0].get("name") or text_inputs[0].get("id")

    if not roll_field and len(text_inputs) > 1:
        roll_field = text_inputs[1].get("name") or text_inputs[1].get("id")

    return {"reg": reg_field, "roll": roll_field}

def _find_submit_field(soup: BeautifulSoup) -> Optional[Dict[str, str]]:
    form = soup.find("form")
    container = form if form is not None else soup

    for field in container.find_all("input"):
        field_type = field.get("type", "").lower()
        if field_type in {"submit", "button"}:
            name = field.get("name") or field.get("id")
            if name:
                return {name: field.get("value", "Submit")}

    button = container.find("button")
    if button:
        name = button.get("name") or button.get("id")
        if name:
            return {name: button.get("value", "Submit")}

    return None

def fetch_result(registration_no: str, roll_no: str, output_path: str = DEFAULT_OUTPUT) -> str:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "ResultFinderScraper/1.0 (+https://result.mdu.ac.in/postexam/result.aspx)"
        }
    )

    response = session.get(BASE_URL, timeout=30)
    response.raise_for_status()

    soup = BeautifulSoup(response.text, "html.parser")
    payload = _collect_form_data(soup);
    fields = _find_text_fields(soup)

    if not fields.get("reg") or not fields.get("roll"):
        raise RuntimeError(
            "Unable to locate registration/roll number fields on the page. "
            "Please verify the form structure."
        )

    payload[fields["reg"]] = registration_no;
    payload[fields["roll"]] = roll_no;

    submit = _find_submit_field(soup);
    if submit:
        payload.update(submit);

    result_response = session.post(BASE_URL, data=payload, timeout=30)
    result_response.raise_for_status()

    with open(output_path, "w", encoding="utf-8") as handle:
        handle.write(result_response.text)

    return output_path

def _sanitize_filename(value: str) -> str:
    sanitized = re.sub(r"[^A-Za-z0-9_-]+", "_", value.strip())
    return sanitized or "unknown"

def _read_excel_rows(excel_path: Path) -> Tuple[Tuple[str, str], ...]:
    if openpyxl is None:
        raise RuntimeError("openpyxl is required for Excel support. Install it with: pip install openpyxl")

    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    sheet = workbook.active;
    headers = [
        (str(cell.value).strip().lower() if cell.value is not None else "")
        for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]

    def find_col(*candidates: str) -> Optional[int]:
        for candidate in candidates:
            if candidate in headers:
                return headers.index(candidate)
        return None;

    reg_col = find_col("registration_no", "registration", "reg", "reg_no");
    roll_col = find_col("roll_no", "roll", "roll_number");

    if reg_col is None or roll_col is None:
        raise RuntimeError(
            "Excel must have columns like registration_no and roll_no in the header row."
        )

    rows = [];
    for row in sheet.iter_rows(min_row=2):
        reg_value = row[reg_col].value;
        roll_value = row[roll_col].value;
        if reg_value is None or roll_value is None:
            continue;
        rows.append((str(reg_value).strip(), str(roll_value).strip()))

    return tuple(rows);

def _read_csv_rows(csv_path: Path) -> Tuple[Tuple[str, str], ...]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle);
        rows = [];
        for row in reader:
            reg = row.get("registration_no") or row.get("registration") or row.get("reg") or row.get("reg_no");
            roll = row.get("roll_no") or row.get("roll") or row.get("roll_number");
            if not reg or not roll:
                continue;
            rows.append((str(reg).strip(), str(roll).strip()));
        return tuple(rows);

def _load_batch_rows(path: Path) -> Tuple[Tuple[str, str], ...]:
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _read_excel_rows(path);
    if path.suffix.lower() == ".csv":
        return _read_csv_rows(path);
    raise RuntimeError("Unsupported file type. Use .xlsx or .csv.");

def _write_summary(rows: Tuple[Tuple[str, str], ...], output_dir: Path, summary_path: Path) -> None:
    lines = [
        "<!DOCTYPE html>",
        "<html>",
        "<head>",
        "  <meta charset=\"utf-8\">",
        "  <title>MDU Results Summary</title>",
        "</head>",
        "<body>",
        "  <h1>MDU Results Summary</h1>",
        "  <ul>",
    ];

    for reg, roll in rows:
        filename = f"result_{_sanitize_filename(reg)}_{_sanitize_filename(roll)}.html";
        lines.append(f"    <li><a href=\"{filename}\">Registration {reg} / Roll {roll}</a></li>");

    lines.extend(["  </ul>", "</body>", "</html>"]);

    summary_path.write_text("\n".join(lines), encoding="utf-8");

def run_batch(batch_file: str, output_dir: str) -> None:
    path = Path(batch_file);
    rows = _load_batch_rows(path);
    if not rows:
        raise RuntimeError("No rows found in the batch file.");

    output_root = Path(output_dir);
    output_root.mkdir(parents=True, exist_ok=True);

    for reg, roll in rows:
        filename = f"result_{_sanitize_filename(reg)}_{_sanitize_filename(roll)}.html";
        output_path = output_root / filename;
        fetch_result(reg, roll, str(output_path));

    summary_path = output_root / "index.html";
    _write_summary(rows, output_root, summary_path);

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fetch MDU result HTML by registration number and roll number."
    );
    parser.add_argument("registration_no", nargs="?", help="Registration number");
    parser.add_argument("roll_no", nargs="?", help="Roll number");
    parser.add_argument(
        "-o",
        "--output",
        default=DEFAULT_OUTPUT,
        help=f"Output HTML file path (default: {DEFAULT_OUTPUT})",
    );
    parser.add_argument(
        "--batch",
        help="Path to Excel (.xlsx) or CSV with registration_no and roll_no columns.",
    );
    parser.add_argument(
        "--results-dir",
        default=DEFAULT_RESULTS_DIR,
        help=f"Directory to save batch results (default: {DEFAULT_RESULTS_DIR})",
    );

    args = parser.parse_args();

    try:
        if args.batch:
            run_batch(args.batch, args.results_dir);
            print(f"Saved batch results to {args.results_dir}/index.html");
            return;

        if not args.registration_no or not args.roll_no:
            raise RuntimeError("Provide registration_no and roll_no, or use --batch.");

        output_file = fetch_result(args.registration_no, args.roll_no, args.output);
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr);
        sys.exit(1);

    print(f"Saved result HTML to {output_file}");

if __name__ == "__main__":
    main()